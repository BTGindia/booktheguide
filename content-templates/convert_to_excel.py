import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# Brand colors
HEADER_FILL = PatternFill(start_color="1A1A18", end_color="1A1A18", fill_type="solid")  # btg-dark
HEADER_FONT = Font(name="Calibri", bold=True, color="F5F0E8", size=11)  # btg-cream text
SECTION_FILL = PatternFill(start_color="58BDAE", end_color="58BDAE", fill_type="solid")  # btg-primary/teal
SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CRITICAL_FILL = PatternFill(start_color="FFDDD2", end_color="FFDDD2", fill_type="solid")  # light coral
HIGH_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # light yellow
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

CSV_FILES = [
    ("01_STATE_HUBS_TEMPLATE.csv", "State Hubs"),
    ("02_CATEGORY_LANDING_TEMPLATE.csv", "Category Landing"),
    ("03_STATE_CATEGORY_TEMPLATE.csv", "State + Category"),
    ("04_GUIDE_PRODUCT_TEMPLATE.csv", "Guide Product"),
    ("05_BLOG_TEMPLATE.csv", "Blog"),
    ("06_UTILITY_PAGES_TEMPLATE.csv", "Utility Pages"),
]


def style_worksheet(ws, rows):
    # Write data
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = NORMAL_FONT

    # Style header row
    for c_idx in range(1, len(rows[0]) + 1 if rows else 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Style data rows based on SEO importance & section changes
    prev_section = None
    for r_idx, row in enumerate(rows[1:], 2):  # skip header
        section = row[0] if len(row) > 0 else ""
        seo_importance = row[4] if len(row) > 4 else ""

        # Section header rows (separator rows like "=== ABOUT PAGE ===")
        if section.startswith("===") or section.startswith("---"):
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.fill = SECTION_FILL
                cell.font = SECTION_FONT
            continue

        # Highlight by SEO importance
        if seo_importance == "Critical":
            for c_idx in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = CRITICAL_FILL
        elif seo_importance == "High":
            for c_idx in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = HIGH_FILL

        # Bold the section name when it changes
        if section and section != prev_section:
            ws.cell(row=r_idx, column=1).font = BOLD_FONT
        prev_section = section

    # Auto-width columns (capped)
    for c_idx in range(1, (len(rows[0]) + 1) if rows else 1):
        max_len = 0
        for row in rows:
            if c_idx <= len(row):
                max_len = max(max_len, len(str(row[c_idx - 1] or "")))
        adjusted = min(max_len + 4, 50)
        ws.column_dimensions[get_column_letter(c_idx)].width = max(adjusted, 12)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"


def convert():
    # --- Individual Excel files ---
    for csv_file, sheet_name in CSV_FILES:
        csv_path = os.path.join(OUTPUT_DIR, csv_file)
        if not os.path.exists(csv_path):
            print(f"  Skipping {csv_file} (not found)")
            continue

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        style_worksheet(ws, rows)

        xlsx_name = csv_file.replace(".csv", ".xlsx")
        xlsx_path = os.path.join(OUTPUT_DIR, xlsx_name)
        wb.save(xlsx_path)
        print(f"  Created: {xlsx_name}")

    # --- Combined workbook with all sheets ---
    wb_all = Workbook()
    wb_all.remove(wb_all.active)  # remove default sheet

    for csv_file, sheet_name in CSV_FILES:
        csv_path = os.path.join(OUTPUT_DIR, csv_file)
        if not os.path.exists(csv_path):
            continue

        ws = wb_all.create_sheet(title=sheet_name)
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        style_worksheet(ws, rows)

    combined_path = os.path.join(OUTPUT_DIR, "BTG_ALL_CONTENT_TEMPLATES.xlsx")
    wb_all.save(combined_path)
    print(f"\n  Combined workbook: BTG_ALL_CONTENT_TEMPLATES.xlsx")
    print(f"\n  All files saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    print("Converting CSV templates to Excel...\n")
    convert()
    print("\nDone!")
